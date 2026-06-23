"""Experiment export functionality for CSV/XLSX download."""

from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import TYPE_CHECKING

from common.logging import get_logger
from contracts.schema_enums import normalize_e_intra_method
from features.common import run_in_session
from features.experiments.e_intra_method import resolve_experiment_e_intra_method

if TYPE_CHECKING:
    from sqlalchemy.orm import Session

logger = get_logger("features.experiments.export")


def is_xlsx_available() -> bool:
    """Check if XLSX export is available (openpyxl installed)."""
    try:
        import openpyxl  # noqa: F401

        return True
    except ImportError:
        return False


# Export columns
EXPORT_COLUMNS = [
    "exp_id",
    "status",
    "run_tier",
    "ff_type",
    "temperature_k",
    "pressure_atm",
    "seed",
    "comp_asphaltene_wt",
    "comp_resin_wt",
    "comp_aromatic_wt",
    "comp_saturate_wt",
    "additive_type",
    "additive_wt",
    "additive_mol_id",
    "e_intra_method",
    "e_intra_method_origin",
    "e_intra_method_resolved_from",
    "e_intra_method_source",
    "wall_time_seconds",
    "created_at",
    "completed_at",
    "density",
    "cohesive_energy_density",
    "viscosity",
]


def _format_datetime(dt: datetime | None) -> str:
    """Format datetime for export."""
    if dt is None:
        return ""
    return dt.isoformat()


def _format_value(value) -> str:
    """Format value for export."""
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.6g}"
    if isinstance(value, datetime):
        return _format_datetime(value)
    return str(value)


def _get_metric_value(exp, metric_name: str) -> float | None:
    """Get metric value from experiment metrics relationship."""
    if not hasattr(exp, "metrics") or not exp.metrics:
        return None
    for metric in exp.metrics:
        if metric.metric_name == metric_name:
            return metric.value
    return None


def _query_experiments(
    session: Session,
    *,
    status: str | None = None,
    tier: str | None = None,
    study_type: str | None = None,
    additive_mol_id: str | None = None,
    e_intra_method: str | None = None,
    limit: int = 1000,
) -> list:
    """Query experiments with optional filters."""
    from sqlalchemy import or_

    from database.models import ExperimentModel

    query = session.query(ExperimentModel)

    if status:
        query = query.filter(ExperimentModel.status == status)
    if tier:
        query = query.filter(ExperimentModel.run_tier == tier)
    if study_type:
        if study_type == "bulk":
            query = query.filter(
                or_(
                    ExperimentModel.study_type == "bulk",
                    ExperimentModel.study_type.is_(None),
                )
            )
        else:
            query = query.filter(ExperimentModel.study_type == study_type)
    if additive_mol_id:
        query = query.filter(ExperimentModel.additive_mol_id == additive_mol_id)

    experiments = query.order_by(ExperimentModel.created_at.desc()).all()
    method_filter = normalize_e_intra_method(e_intra_method)
    if method_filter:
        experiments = [
            exp for exp in experiments if resolve_experiment_e_intra_method(exp)[0] == method_filter
        ]
    return experiments[:limit]


def _experiment_to_row(exp) -> dict:
    """Convert experiment model to export row dict."""
    (
        e_intra_method,
        e_intra_method_origin,
        e_intra_method_resolved_from,
    ) = resolve_experiment_e_intra_method(exp)
    return {
        "exp_id": exp.exp_id,
        "status": exp.status,
        "run_tier": exp.run_tier,
        "ff_type": exp.ff_type,
        "temperature_k": exp.temperature_K,
        "pressure_atm": exp.pressure_atm,
        "seed": exp.seed,
        "comp_asphaltene_wt": exp.comp_asphaltene_wt,
        "comp_resin_wt": exp.comp_resin_wt,
        "comp_aromatic_wt": exp.comp_aromatic_wt,
        "comp_saturate_wt": exp.comp_saturate_wt,
        "additive_type": exp.additive_type,
        "additive_wt": exp.additive_wt,
        "additive_mol_id": exp.additive_mol_id,
        "e_intra_method": e_intra_method,
        "e_intra_method_origin": e_intra_method_origin,
        "e_intra_method_resolved_from": e_intra_method_resolved_from,
        "e_intra_method_source": e_intra_method_origin,
        "wall_time_seconds": exp.wall_time_seconds,
        "created_at": exp.created_at,
        "completed_at": exp.completed_at,
        "density": _get_metric_value(exp, "density"),
        "cohesive_energy_density": _get_metric_value(exp, "cohesive_energy_density"),
        "viscosity": _get_metric_value(exp, "viscosity"),
    }


def export_experiments_csv(
    *,
    status: str | None = None,
    tier: str | None = None,
    study_type: str | None = None,
    additive_mol_id: str | None = None,
    e_intra_method: str | None = None,
    limit: int = 1000,
) -> str:
    """Export experiments to CSV string.

    Args:
        status: Optional status filter
        tier: Optional tier filter
        limit: Maximum number of experiments

    Returns:
        CSV content as string
    """

    def _export(session: Session) -> str:
        experiments = _query_experiments(
            session,
            status=status,
            tier=tier,
            study_type=study_type,
            additive_mol_id=additive_mol_id,
            e_intra_method=e_intra_method,
            limit=limit,
        )

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=EXPORT_COLUMNS)
        writer.writeheader()

        for exp in experiments:
            row = _experiment_to_row(exp)
            formatted_row = {k: _format_value(row.get(k)) for k in EXPORT_COLUMNS}
            writer.writerow(formatted_row)

        return output.getvalue()

    return run_in_session(_export)


def export_experiments_xlsx(
    *,
    status: str | None = None,
    tier: str | None = None,
    study_type: str | None = None,
    additive_mol_id: str | None = None,
    e_intra_method: str | None = None,
    limit: int = 1000,
) -> bytes:
    """Export experiments to XLSX bytes.

    Args:
        status: Optional status filter
        tier: Optional tier filter
        limit: Maximum number of experiments

    Returns:
        XLSX content as bytes

    Raises:
        ImportError: If openpyxl is not installed
    """
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for XLSX export. Install with: pip install openpyxl"
        ) from exc

    def _export(session: Session) -> bytes:
        experiments = _query_experiments(
            session,
            status=status,
            tier=tier,
            study_type=study_type,
            additive_mol_id=additive_mol_id,
            e_intra_method=e_intra_method,
            limit=limit,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Experiments"

        # Header row
        for col_idx, col_name in enumerate(EXPORT_COLUMNS, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Data rows
        for row_idx, exp in enumerate(experiments, start=2):
            row = _experiment_to_row(exp)
            for col_idx, col_name in enumerate(EXPORT_COLUMNS, start=1):
                value = row.get(col_name)
                if isinstance(value, datetime):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                elif value is not None:
                    ws.cell(row=row_idx, column=col_idx, value=value)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    return run_in_session(_export)
